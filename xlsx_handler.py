# Ten moduł odpowiedzialny jest za podstawowe operacje wykonywane w plikach Excel w projekcie, takich jak Próbkowanie i Łączenie wyników.
from xlsx_template import XlsxWzor
from samples import Probkowanie
import pandas as pd

# OpenPyxl Shenenigans
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill

class XlsxHandler:
    def __init__(self) -> None:
        
        self.probki: list[str] = []
        self.sciezka: str = ""
        self.content: pd.DataFrame | None = None 
        
    def otwórz_plik(self, ścieżka: str) -> None:
        """Otwiera plik do analizy"""

        self.probki = {}
        self.sciezka = ścieżka
        self.content = pd.read_excel(ścieżka)

        # Rename the column "Unnamed: 0" to an empty string
        self.content.rename(columns={"Unnamed: 0": ""}, inplace=True)
        
        # Check if columns match with the template
        if self.content.columns.tolist() != XlsxWzor.wez_wszystkie_kolumny():
            raise TypeError(f"Uwaga! Kolumny pliku: {self.sciezka} nie pokrywają się z template'em!")

    def zmodyfikuj_zawartość_o_daty(self, daty: list[str]) -> None:
        """Modyfikuje zawartość pliku, jeśli chodzi o daty do wyboru."""

        # Jeśli nie ma dat:
        if not daty:
            return

        daty = [pd.to_datetime(data, format='%d.%m.%Y') for data in daty]
        if daty[0] != daty[1]:
            self.content = self.content.loc[self.content[XlsxWzor.DATA_URUCHOMIENIA].between(daty[0], daty[1])]
        else:
            self.content = self.content.loc[self.content[XlsxWzor.DATA_URUCHOMIENIA] == pd.to_datetime(daty[0], format='%d.%m.%Y')]

    def przeczytaj_możliwe_daty(self) -> list[str]:
        """Bierze z pliku do tablicy wszystkie daty z kolumny data uruchomienia."""
        
        return [data.strftime("%d.%m.%Y") for data in self.content[XlsxWzor.DATA_URUCHOMIENIA].tolist()]
 
    def zapisz_wybrane_próbki(self) -> pd.DataFrame:
        """Zapisuje wybrane już próbki razem z plikiem, z którego pochodzą."""

        frame: pd.DataFrame = self.content.loc[self.content[XlsxWzor.NUMER_WNIOSKU].isin(self.probki)].copy()

        if len(frame) == 0:
            print(f"UWAGA! Nie ma żadnych pasujących elementów w pliku {self.sciezka}")
            return 

        frame.loc[:, XlsxWzor.PLIK_ŹRÓDŁOWY_ŚCIEŻKA] = self.sciezka
        frame.loc[:, XlsxWzor.PLIK_ŹRÓDŁOWY_NAZWA] = self.sciezka.split('\\')[-1]
        
        najważniejsze_kolumny: list[str] = XlsxWzor.wez_najważniejsze_kolumny()
        
        połączony_frame: pd.DataFrame = pd.concat([
            frame[[XlsxWzor.PLIK_ŹRÓDŁOWY_NAZWA]],
            frame[najważniejsze_kolumny]
        ], axis=1, ignore_index=True, sort=False)

        połączony_frame.columns = [
            XlsxWzor.PLIK_ŹRÓDŁOWY_NAZWA
        ] + najważniejsze_kolumny
        return połączony_frame

    def wylosuj_próbki(self) -> None:
        """Losuje próbki do pokolorowania. """

        if len(self.content) == 0:
            return
        
        probkowanie = Probkowanie(self.content)
        probkowanie.wylosuj_próbki()
        # Wypłaszcza dict z wybranymi próbkami.
        self.probki = [i for k in probkowanie.próbki.values() for i in k]

    def pokoloruj_plik(self) -> None:
        """Zajmuje się kolorowaniem pliku na podstawie wybranych próbek."""

        skoroszyt: Workbook = load_workbook(self.sciezka, keep_vba=True)
        arkusz: Worksheet = skoroszyt.worksheets[0]
        
        # Tutaj można zmienić kolor - text FFFF00 oznacza żółty w HEX DEC.
        color: PatternFill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for r in range(2, arkusz.max_row +1):
            komórka: Cell = arkusz.cell(int(r), self.content.columns.to_list().index(XlsxWzor.NUMER_WNIOSKU)+1)
            if komórka.value in self.probki:
                komórka.fill = color
            else:
                komórka.style = "Normal"
        skoroszyt.save(self.sciezka)
